import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def getColumnCount(file,sheetname):
    worbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnnum).value

def writeData(file,sheetname,rownum,columnnum,value):
    orkbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=columnnum).value=data
    workbook.save(file)