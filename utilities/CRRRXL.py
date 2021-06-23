import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def getRowCount(file,sheetname,rownum,colunmnum):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum=row,colunmnum=column).value

def getRowCount(file,sheetname,rownum,colunmnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum=row,colunmnum=column).value=data
    workbook.save(file)